"""
Av1Automóveis — Gerenciador de Estoque
Gerencia carros.xlsx e exporta carros.json automaticamente para o GitHub Pages.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys, shutil, json
from pathlib import Path
from datetime import datetime

BASE_DIR   = Path(__file__).parent.parent
DATA_FILE  = BASE_DIR / "data" / "carros.xlsx"
JSON_FILE  = BASE_DIR / "carros.json"        # lido pelo GitHub Pages
IMAGES_DIR = BASE_DIR / "images"
IMAGES_DIR.mkdir(exist_ok=True)

HEADERS = ["id","nome","marca","ano","preco","quilometragem",
           "potencia","combustivel","cambio","descricao","imagem_path","status"]

MARCAS = sorted(["Acura","Alfa Romeo","Audi","BMW","BYD","Caoa Chery","Chevrolet",
          "Chrysler","Citroën","Dodge","Ferrari","Fiat","Ford","GWM","Honda",
          "Hyundai","Infiniti","Jaguar","Jeep","Kia","Lamborghini","Land Rover",
          "Lexus","Maserati","Mazda","Mercedes-Benz","Mitsubishi","Nissan",
          "Peugeot","Porsche","RAM","Renault","Subaru","Suzuki","Toyota",
          "Volkswagen","Volvo","Outra"])

COMBUSTIVEIS = ["Flex","Gasolina","Etanol","Diesel","Elétrico","Híbrido","GNV"]
CAMBIOS      = ["Manual","Automático","CVT","Automatizado","Dupla Embreagem"]
STATUS_OPTS  = ["disponivel","vendido","reservado"]

C = {
    "bg":"#F5F7FA","sidebar":"#1C2340","sidebar2":"#141A30",
    "accent":"#1C2340","accent2":"#E8C840","white":"#FFFFFF",
    "text":"#1C2340","gray":"#6B7280","border":"#E5E7EB",
    "success":"#10B981","warning":"#F59E0B","danger":"#EF4444",
    "row_odd":"#FFFFFF","row_even":"#F9FAFB",
}

# ── Excel ────────────────────────────────────────────────────────────────────
def load_rows():
    if not DATA_FILE.exists():
        _init_workbook()
    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(c is not None for c in row):
            rows.append(list(row))
    return rows

def save_rows(rows):
    if not DATA_FILE.exists():
        _init_workbook()
    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    dfont  = Font(size=10)
    center = Alignment(horizontal="center", vertical="center")
    for ri, rd in enumerate(rows, start=2):
        for ci, val in enumerate(rd, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = dfont; cell.border = border
            if ci in [1,3,4,5,6,7,8,9,12]:
                cell.alignment = center
    wb.save(DATA_FILE)

def _init_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Carros"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [6,25,15,8,15,15,12,14,12,50,30,12]
    for col, (h, w) in enumerate(zip(HEADERS, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="1a1a2e")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    DATA_FILE.parent.mkdir(exist_ok=True)
    wb.save(DATA_FILE)

def next_id(rows):
    ids = [r[0] for r in rows if r[0] is not None]
    return max(ids) + 1 if ids else 1

def copy_image(src, car_id):
    ext  = Path(src).suffix.lower()
    dest = IMAGES_DIR / f"carro_{car_id}{ext}"
    shutil.copy2(src, dest)
    return f"images/carro_{car_id}{ext}"

# ── JSON export ──────────────────────────────────────────────────────────────
def export_json(rows=None):
    if rows is None:
        rows = load_rows()
    cars = []
    for r in rows:
        if r[0] is None:
            continue
        cars.append({
            "id":            int(r[0]),
            "nome":          str(r[1] or ""),
            "marca":         str(r[2] or ""),
            "ano":           int(r[3]) if r[3] else None,
            "preco":         float(r[4]) if r[4] else 0,
            "quilometragem": float(r[5]) if r[5] else 0,
            "potencia":      str(r[6] or ""),
            "combustivel":   str(r[7] or ""),
            "cambio":        str(r[8] or ""),
            "descricao":     str(r[9] or ""),
            "imagem_path":   str(r[10] or ""),
            "status":        str(r[11] or "disponivel"),
        })
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(cars, f, ensure_ascii=False, indent=2)
    return len(cars)


# ── APP ──────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Av1Automóveis — Gerenciador")
        self.geometry("1200x740")
        self.minsize(960, 600)
        self.configure(bg=C["bg"])
        self.selected_img = tk.StringVar()
        self.edit_id = None
        self._build_ui()
        self.refresh_table()

    # ── Layout ───────────────────────────────────────────────────────────────
    def _build_ui(self):
        self.sidebar = tk.Frame(self, bg=C["sidebar"], width=230)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        self.sidebar.pack_propagate(False)
        self.main = tk.Frame(self, bg=C["bg"])
        self.main.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._build_sidebar()
        self._build_main()

    def _build_sidebar(self):
        lf = tk.Frame(self.sidebar, bg=C["sidebar"], pady=28)
        lf.pack(fill=tk.X)
        tk.Label(lf, text="AV.1", font=("Helvetica",26,"bold"),
                 bg=C["sidebar"], fg=C["accent2"]).pack()
        tk.Label(lf, text="AUTOMÓVEIS", font=("Helvetica",9,"bold"),
                 bg=C["sidebar"], fg="#8892A4").pack()
        tk.Frame(self.sidebar, bg="#2D3748", height=1).pack(fill=tk.X, padx=20, pady=4)

        for label, cmd in [
            ("🚗  Estoque",         self.show_list),
            ("➕  Adicionar Carro", self.show_add),
            ("📊  Resumo",          self.show_summary),
        ]:
            self._nav_btn(label, cmd)

        # Export button (prominent)
        tk.Frame(self.sidebar, bg="#2D3748", height=1).pack(fill=tk.X, padx=20, pady=12)
        export_btn = tk.Button(
            self.sidebar, text="⬆  Exportar JSON\npara GitHub Pages",
            font=("Helvetica",9,"bold"), bd=0, relief=tk.FLAT,
            bg=C["accent2"], fg="white", cursor="hand2",
            padx=20, pady=14, command=self._do_export, justify=tk.LEFT)
        export_btn.pack(fill=tk.X, padx=14, pady=4)
        export_btn.bind("<Enter>", lambda e: export_btn.config(bg="#e8192c"))
        export_btn.bind("<Leave>", lambda e: export_btn.config(bg=C["accent2"]))

        self.export_lbl = tk.Label(self.sidebar, text="",
                                   font=("Helvetica",8), bg=C["sidebar"], fg="#6B7280",
                                   wraplength=190, justify=tk.LEFT)
        self.export_lbl.pack(padx=14, fill=tk.X)

        tk.Label(self.sidebar, text="v2.0 · GitHub Pages",
                 font=("Helvetica",8), bg=C["sidebar"],
                 fg="#4A5568").pack(side=tk.BOTTOM, pady=12)

    def _nav_btn(self, text, cmd):
        btn = tk.Button(self.sidebar, text=text, anchor="w",
                        font=("Helvetica",10), bd=0, relief=tk.FLAT,
                        bg=C["sidebar"], fg="#CBD5E0",
                        activebackground=C["accent"], activeforeground="white",
                        cursor="hand2", padx=20, pady=12, command=cmd)
        btn.pack(fill=tk.X)
        btn.bind("<Enter>", lambda e: btn.config(bg=C["accent"]))
        btn.bind("<Leave>", lambda e: btn.config(bg=C["sidebar"]))

    def _build_main(self):
        self.header = tk.Frame(self.main, bg=C["white"], height=64)
        self.header.pack(fill=tk.X)
        self.header.pack_propagate(False)
        self.page_title = tk.Label(self.header, text="Estoque",
                                   font=("Helvetica",16,"bold"),
                                   bg=C["white"], fg=C["text"])
        self.page_title.pack(side=tk.LEFT, padx=24, pady=16)
        self.page_sub = tk.Label(self.header, text="",
                                 font=("Helvetica",10),
                                 bg=C["white"], fg=C["gray"])
        self.page_sub.pack(side=tk.LEFT, pady=16)

        self.content = tk.Frame(self.main, bg=C["bg"])
        self.content.pack(fill=tk.BOTH, expand=True)

        self._build_list_frame()
        self._build_form_frame()
        self._build_summary_frame()

    # ── Export ───────────────────────────────────────────────────────────────
    def _do_export(self, silent=False):
        try:
            n = export_json()
            ts = datetime.now().strftime("%H:%M:%S")
            self.export_lbl.config(
                text=f"✔ {n} carros exportados\nÚltimo: {ts}",
                fg=C["success"])
            if not silent:
                messagebox.showinfo(
                    "JSON Exportado!",
                    f"carros.json gerado com {n} veículos.\n\n"
                    "Próximo passo:\n"
                    "Faça commit e push do arquivo carros.json\n"
                    "(e das imagens novas em /images)\n"
                    "para o repositório do GitHub.")
        except Exception as ex:
            messagebox.showerror("Erro ao exportar", str(ex))

    # ── List ─────────────────────────────────────────────────────────────────
    def _build_list_frame(self):
        self.list_frame = tk.Frame(self.content, bg=C["bg"])

        bar = tk.Frame(self.list_frame, bg=C["bg"], pady=12)
        bar.pack(fill=tk.X, padx=20)
        tk.Label(bar, text="Buscar:", font=("Helvetica",10),
                 bg=C["bg"], fg=C["text"]).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *a: self.refresh_table())
        e = tk.Entry(bar, textvariable=self.search_var, font=("Helvetica",10),
                     width=26, relief=tk.FLAT, bd=0,
                     highlightthickness=1, highlightbackground=C["border"],
                     highlightcolor=C["accent"], bg=C["white"])
        e.pack(side=tk.LEFT, padx=(4,20), ipady=6, ipadx=6)

        tk.Label(bar, text="Status:", font=("Helvetica",10),
                 bg=C["bg"], fg=C["text"]).pack(side=tk.LEFT)
        self.filter_st = ttk.Combobox(bar, values=["Todos"]+STATUS_OPTS,
                                      state="readonly", width=14,
                                      font=("Helvetica",10))
        self.filter_st.set("Todos")
        self.filter_st.bind("<<ComboboxSelected>>", lambda e: self.refresh_table())
        self.filter_st.pack(side=tk.LEFT, padx=(4,0), ipady=4)

        tf = tk.Frame(self.list_frame, bg=C["white"],
                      highlightthickness=1, highlightbackground=C["border"])
        tf.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0,12))

        cols = ("ID","Nome","Marca","Ano","Preço","Km","Combustível","Câmbio","Status")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", selectmode="browse")
        cw = {"ID":45,"Nome":210,"Marca":90,"Ano":55,"Preço":110,
              "Km":90,"Combustível":90,"Câmbio":100,"Status":90}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=cw[c],
                             anchor="w" if c == "Nome" else "center")

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._style_tree()

        af = tk.Frame(self.list_frame, bg=C["bg"], pady=8)
        af.pack(fill=tk.X, padx=20)
        self._abtn(af,"✏️  Editar",      C["accent"],  self._edit_sel).pack(side=tk.LEFT, padx=(0,8))
        self._abtn(af,"✅  Vendido",     C["success"], lambda: self._chg_status("vendido")).pack(side=tk.LEFT, padx=(0,8))
        self._abtn(af,"⏳  Reservar",    C["warning"], lambda: self._chg_status("reservado")).pack(side=tk.LEFT, padx=(0,8))
        self._abtn(af,"🔄  Disponível",  "#6366F1",    lambda: self._chg_status("disponivel")).pack(side=tk.LEFT, padx=(0,8))
        self._abtn(af,"🗑️  Remover",     C["danger"],  self._remove_sel).pack(side=tk.RIGHT)

    def _style_tree(self):
        s = ttk.Style(); s.theme_use("clam")
        s.configure("Treeview", font=("Helvetica",10), rowheight=34,
                    background=C["white"], fieldbackground=C["white"],
                    foreground=C["text"], borderwidth=0)
        s.configure("Treeview.Heading", font=("Helvetica",10,"bold"),
                    background=C["sidebar"], foreground="white",
                    relief="flat", padding=(0,8))
        s.map("Treeview",
              background=[("selected", C["accent2"])],
              foreground=[("selected", C["sidebar"])])
        self.tree.tag_configure("disponivel", background=C["row_odd"])
        self.tree.tag_configure("vendido",    background="#FEF2F2", foreground="#DC2626")
        self.tree.tag_configure("reservado",  background="#FFFBEB", foreground="#D97706")
        self.tree.tag_configure("even",       background=C["row_even"])

    def _abtn(self, parent, text, color, command):
        btn = tk.Button(parent, text=text, font=("Helvetica",9,"bold"),
                        bg=color, fg="white", bd=0, relief=tk.FLAT,
                        cursor="hand2", padx=14, pady=7, command=command)
        r = lambda c: max(0, int(c[1:3],16)-20)
        g = lambda c: max(0, int(c[3:5],16)-20)
        b = lambda c: max(0, int(c[5:7],16)-20)
        dk = f"#{r(color):02x}{g(color):02x}{b(color):02x}"
        btn.bind("<Enter>", lambda e: btn.config(bg=dk))
        btn.bind("<Leave>", lambda e: btn.config(bg=color))
        return btn

    def refresh_table(self, *_):
        for item in self.tree.get_children():
            self.tree.delete(item)
        rows = load_rows()
        search = self.search_var.get().lower()
        fst    = self.filter_st.get()
        filtered = []
        for r in rows:
            if not r[0]: continue
            if search and search not in str(r[1]or"").lower() and search not in str(r[2]or"").lower():
                continue
            st = str(r[11] or "disponivel")
            if fst != "Todos" and st != fst: continue
            filtered.append(r)
        for i, r in enumerate(filtered):
            st = str(r[11] or "disponivel")
            preco = f"R$ {float(r[4] or 0):,.0f}".replace(",",".")
            km    = f"{int(float(r[5] or 0)):,} km".replace(",",".")
            tag   = st if st in ("vendido","reservado","disponivel") else ("even" if i%2 else "disponivel")
            self.tree.insert("", "end", iid=str(r[0]), tags=(tag,),
                             values=(r[0],r[1],r[2],r[3],preco,km,r[7],r[8],st.capitalize()))
        self.page_sub.config(text=f"({len(filtered)} veículo{'s' if len(filtered)!=1 else ''})")

    def _get_sel(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atenção","Selecione um veículo.")
            return None
        cid = int(sel[0])
        for r in load_rows():
            if r[0] == cid: return r
        return None

    def _chg_status(self, st):
        row = self._get_sel()
        if not row: return
        rows = load_rows()
        for r in rows:
            if r[0] == row[0]: r[11] = st
        save_rows(rows)
        export_json(rows)
        self.export_lbl.config(text=f"✔ JSON atualizado\n{datetime.now().strftime('%H:%M:%S')}", fg=C["success"])
        self.refresh_table()

    def _remove_sel(self):
        row = self._get_sel()
        if not row: return
        if not messagebox.askyesno("Confirmar", f"Remover '{row[1]}'?\nEsta ação não pode ser desfeita."):
            return
        rows = [r for r in load_rows() if r[0] != row[0]]
        save_rows(rows)
        export_json(rows)
        self.export_lbl.config(text=f"✔ JSON atualizado\n{datetime.now().strftime('%H:%M:%S')}", fg=C["success"])
        self.refresh_table()
        messagebox.showinfo("Removido","Veículo removido e JSON atualizado.")

    def _edit_sel(self):
        row = self._get_sel()
        if not row: return
        self.edit_id = row[0]
        self.show_add(edit_data=row)

    # ── Form ─────────────────────────────────────────────────────────────────
    def _build_form_frame(self):
        self.form_frame = tk.Frame(self.content, bg=C["bg"])
        canvas = tk.Canvas(self.form_frame, bg=C["bg"], highlightthickness=0)
        vsb = ttk.Scrollbar(self.form_frame, orient="vertical", command=canvas.yview)
        self.form_inner = tk.Frame(canvas, bg=C["bg"])
        self.form_inner.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=self.form_inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)),"units"))
        self._build_fields()

    def _build_fields(self):
        p = self.form_inner; pad = {"padx":24,"pady":5}

        def sec(txt):
            f = tk.Frame(p, bg=C["bg"]); f.pack(fill=tk.X, **pad)
            tk.Label(f, text=txt, font=("Helvetica",11,"bold"),
                     bg=C["bg"], fg=C["accent"]).pack(side=tk.LEFT)
            tk.Frame(f, bg=C["border"], height=1).pack(
                side=tk.LEFT, fill=tk.X, expand=True, padx=(10,0))

        def row(): f = tk.Frame(p, bg=C["bg"]); f.pack(fill=tk.X, padx=24, pady=3); return f

        def fld(parent, label):
            lf = tk.Frame(parent, bg=C["bg"])
            lf.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0,14))
            tk.Label(lf, text=label, font=("Helvetica",9),
                     bg=C["bg"], fg=C["gray"]).pack(anchor="w")
            return lf

        def ent(parent, var):
            e = tk.Entry(parent, textvariable=var, font=("Helvetica",10),
                         relief=tk.FLAT, bd=0, highlightthickness=1,
                         highlightbackground=C["border"], highlightcolor=C["accent"],
                         bg=C["white"])
            e.pack(fill=tk.X, ipady=7, ipadx=6); return e

        def cmb(parent, var, vals):
            cb = ttk.Combobox(parent, textvariable=var, values=vals,
                              state="readonly", font=("Helvetica",10))
            cb.pack(fill=tk.X, ipady=4); return cb

        self.f = {k: tk.StringVar() for k in
                  ["nome","marca","ano","preco","quilometragem","potencia",
                   "combustivel","cambio","status"]}

        sec("Informações Básicas")
        r = row()
        ent(fld(r,"Nome / Modelo *"), self.f["nome"])
        cmb(fld(r,"Marca *"), self.f["marca"], MARCAS)

        r2 = row()
        ent(fld(r2,"Ano *"),      self.f["ano"])
        ent(fld(r2,"Preço (R$) *"), self.f["preco"])
        cmb(fld(r2,"Status *"), self.f["status"], STATUS_OPTS)

        sec("Especificações")
        r3 = row()
        ent(fld(r3,"Quilometragem (km) *"), self.f["quilometragem"])
        ent(fld(r3,"Potência (ex: 177cv)"), self.f["potencia"])
        r4 = row()
        cmb(fld(r4,"Combustível *"), self.f["combustivel"], COMBUSTIVEIS)
        cmb(fld(r4,"Câmbio *"), self.f["cambio"], CAMBIOS)

        sec("Descrição")
        df = tk.Frame(p, bg=C["bg"]); df.pack(fill=tk.X, padx=24, pady=3)
        tk.Label(df, text="Descrição do veículo",
                 font=("Helvetica",9), bg=C["bg"], fg=C["gray"]).pack(anchor="w")
        self.f_desc = tk.Text(df, height=5, font=("Helvetica",10),
                              relief=tk.FLAT, bd=0, highlightthickness=1,
                              highlightbackground=C["border"],
                              highlightcolor=C["accent"], bg=C["white"],
                              padx=6, pady=6, wrap=tk.WORD)
        self.f_desc.pack(fill=tk.X)

        sec("Imagem")
        imgf = tk.Frame(p, bg=C["bg"]); imgf.pack(fill=tk.X, padx=24, pady=3)
        self.img_lbl = tk.Label(imgf, text="Nenhuma imagem selecionada",
                                font=("Helvetica",9), bg=C["bg"], fg=C["gray"])
        self.img_lbl.pack(side=tk.LEFT)
        self._abtn(imgf,"📁  Selecionar Imagem", C["accent"],
                   self._pick_img).pack(side=tk.LEFT, padx=(12,0))

        bf = tk.Frame(p, bg=C["bg"]); bf.pack(fill=tk.X, padx=24, pady=(18,28))
        self.save_btn = self._abtn(bf,"💾  Salvar e Exportar JSON", C["success"], self._save)
        self.save_btn.pack(side=tk.LEFT, padx=(0,12))
        self._abtn(bf,"✖  Cancelar", C["gray"], self.show_list).pack(side=tk.LEFT)

    def _pick_img(self):
        path = filedialog.askopenfilename(
            title="Selecionar Imagem",
            filetypes=[("Imagens","*.jpg *.jpeg *.png *.webp *.bmp")])
        if path:
            self.selected_img.set(path)
            self.img_lbl.config(text=f"✔ {Path(path).name}", fg=C["success"])

    def _save(self):
        required = {"nome":"Nome","marca":"Marca","ano":"Ano","preco":"Preço",
                    "quilometragem":"Quilometragem","combustivel":"Combustível",
                    "cambio":"Câmbio","status":"Status"}
        for k, lbl in required.items():
            if not self.f[k].get().strip():
                messagebox.showerror("Campo obrigatório", f"'{lbl}' é obrigatório."); return
        try:
            ano   = int(self.f["ano"].get())
            preco = float(self.f["preco"].get().replace(".","").replace(",","."))
            km    = float(self.f["quilometragem"].get().replace(".","").replace(",","."))
        except ValueError:
            messagebox.showerror("Erro","Ano, Preço e Quilometragem devem ser números."); return

        rows   = load_rows()
        car_id = self.edit_id if self.edit_id else next_id(rows)

        img_path = ""
        if self.selected_img.get():
            try:
                img_path = copy_image(self.selected_img.get(), car_id)
            except Exception as ex:
                messagebox.showerror("Erro na imagem", str(ex)); return
        elif self.edit_id:
            for r in rows:
                if r[0] == self.edit_id:
                    img_path = r[10] or ""; break

        new_row = [car_id, self.f["nome"].get().strip(),
                   self.f["marca"].get().strip(), ano, preco, km,
                   self.f["potencia"].get().strip(),
                   self.f["combustivel"].get().strip(),
                   self.f["cambio"].get().strip(),
                   self.f_desc.get("1.0",tk.END).strip(),
                   img_path, self.f["status"].get().strip()]

        if self.edit_id:
            rows = [new_row if r[0]==self.edit_id else r for r in rows]
        else:
            rows.append(new_row)

        save_rows(rows)
        n = export_json(rows)
        self.export_lbl.config(
            text=f"✔ {n} carros exportados\n{datetime.now().strftime('%H:%M:%S')}",
            fg=C["success"])
        self.edit_id = None
        messagebox.showinfo("Salvo!",
            f"'{new_row[1]}' salvo!\n\n"
            "carros.json foi atualizado automaticamente.\n"
            "Faça push para o GitHub para publicar no site.")
        self.show_list()

    def show_add(self, edit_data=None):
        self._clear_form()
        if edit_data:
            self.page_title.config(text="Editar Veículo")
            self.f["nome"].set(edit_data[1] or "")
            self.f["marca"].set(edit_data[2] or "")
            self.f["ano"].set(str(edit_data[3] or ""))
            self.f["preco"].set(str(int(float(edit_data[4] or 0))))
            self.f["quilometragem"].set(str(int(float(edit_data[5] or 0))))
            self.f["potencia"].set(edit_data[6] or "")
            self.f["combustivel"].set(edit_data[7] or "")
            self.f["cambio"].set(edit_data[8] or "")
            self.f_desc.insert("1.0", edit_data[9] or "")
            self.f["status"].set(edit_data[11] or "disponivel")
            if edit_data[10]:
                self.img_lbl.config(text=f"✔ {edit_data[10]}", fg=C["success"])
        else:
            self.edit_id = None
            self.page_title.config(text="Adicionar Veículo")
            self.f["status"].set("disponivel")
        self.page_sub.config(text="")
        self._switch(self.form_frame)

    def _clear_form(self):
        for v in self.f.values(): v.set("")
        self.f_desc.delete("1.0", tk.END)
        self.selected_img.set("")
        self.img_lbl.config(text="Nenhuma imagem selecionada", fg=C["gray"])

    # ── Summary ───────────────────────────────────────────────────────────────
    def _build_summary_frame(self):
        self.summary_frame = tk.Frame(self.content, bg=C["bg"])

    def show_summary(self):
        self.page_title.config(text="Resumo do Estoque")
        self.page_sub.config(text="")
        for w in self.summary_frame.winfo_children(): w.destroy()
        rows = load_rows()
        total = len(rows)
        disp  = sum(1 for r in rows if str(r[11])=="disponivel")
        vend  = sum(1 for r in rows if str(r[11])=="vendido")
        res   = sum(1 for r in rows if str(r[11])=="reservado")
        val   = sum(float(r[4] or 0) for r in rows if str(r[11])=="disponivel")

        cf = tk.Frame(self.summary_frame, bg=C["bg"])
        cf.pack(fill=tk.X, padx=24, pady=24)
        for title, v, color, icon in [
            ("Total",       str(total), C["accent"],  "🚗"),
            ("Disponíveis", str(disp),  C["success"], "✅"),
            ("Vendidos",    str(vend),  C["danger"],  "🏁"),
            ("Reservados",  str(res),   C["warning"], "⏳"),
        ]:
            c = tk.Frame(cf, bg=C["white"], highlightthickness=1,
                         highlightbackground=C["border"])
            c.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,12))
            tk.Label(c, text=icon, font=("Helvetica",22), bg=C["white"]).pack(pady=(18,4))
            tk.Label(c, text=v, font=("Helvetica",28,"bold"),
                     bg=C["white"], fg=color).pack()
            tk.Label(c, text=title, font=("Helvetica",9),
                     bg=C["white"], fg=C["gray"]).pack(pady=(2,18))

        vf = tk.Frame(self.summary_frame, bg=C["white"],
                      highlightthickness=1, highlightbackground=C["border"])
        vf.pack(fill=tk.X, padx=24)
        tk.Label(vf, text="💰  Valor total em estoque (disponíveis)",
                 font=("Helvetica",11,"bold"), bg=C["white"], fg=C["text"]
                 ).pack(side=tk.LEFT, padx=20, pady=16)
        formatted = f"R$ {val:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        tk.Label(vf, text=formatted, font=("Helvetica",14,"bold"),
                 bg=C["white"], fg=C["success"]).pack(side=tk.RIGHT, padx=20)

        # JSON status
        jf = tk.Frame(self.summary_frame, bg=C["bg"]); jf.pack(fill=tk.X, padx=24, pady=12)
        json_exists = JSON_FILE.exists()
        ts = datetime.fromtimestamp(JSON_FILE.stat().st_mtime).strftime("%d/%m/%Y %H:%M") if json_exists else "—"
        tk.Label(jf, text=f"📄 carros.json → {'✔ Existe' if json_exists else '✖ Não exportado'} · Atualizado: {ts}",
                 font=("Helvetica",10), bg=C["bg"],
                 fg=C["success"] if json_exists else C["danger"]).pack(side=tk.LEFT)
        self._abtn(jf,"⬆  Exportar Agora", C["accent2"], self._do_export
                   ).pack(side=tk.LEFT, padx=12)

        self._switch(self.summary_frame)

    # ── Nav ───────────────────────────────────────────────────────────────────
    def show_list(self):
        self.edit_id = None
        self.page_title.config(text="Estoque de Veículos")
        self.refresh_table()
        self._switch(self.list_frame)

    def _switch(self, frame):
        for f in (self.list_frame, self.form_frame, self.summary_frame):
            f.pack_forget()
        frame.pack(fill=tk.BOTH, expand=True)


if __name__ == "__main__":
    app = App()
    app.mainloop()
